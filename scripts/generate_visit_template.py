import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADERS = [
    ("store_name", "Nome da Loja"),
    ("city", "Cidade"),
    ("ingredient_id", "Ingrediente (nome canônico)"),
    ("raw_product", "Produto encontrado (texto da etiqueta)"),
    ("raw_price", "Preço (R$)"),
    ("raw_unit", "Unidade (ex: 1kg, 500g, cx 12x395g)"),
    ("visit_date", "Data da visita (AAAA-MM-DD)"),
    ("notes", "Observações"),
]

STORES_SP = [
    "Manos Doces",
    "Dijos Doces",
    "Jabaquara Doces",
    "Nova Paulista Doces",
    "Patikal Doces",
    "Marsil Atacadista",
    "Vito Quintas",
    "Doces Cabo Verde",
    "Fernandez Atacadista",
    "Campineira Doces",
    "Embalagens Cantareira",
]

INGREDIENTS = [
    "Leite Condensado Integral",
    "Creme de Leite 20% Gordura",
    "Chocolate em Pó 50% Cacau",
    "Leite Ninho Integral",
    "Granulado Melken Ao Leite",
    "Granulado Melken Branco",
    "Granulado Melken Meio Amargo",
    "Nutella",
    "Coloretti Granulado Colorido",
    "Coco Ralado Grosso sem Açúcar",
    "Chocolate Nobre Blend Harald",
]


def generate_template(output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Visita Mensal"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, (key, label) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 30

    row = 3
    ws.cell(row=row, column=1, value="INSTRUÇÕES:").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="1. Preencha uma linha por produto encontrado em cada loja visitada.")
    row += 1
    ws.cell(row=row, column=1, value="2. 'Ingrediente' = nome canônico da lista acima (cópia e cola).")
    row += 1
    ws.cell(row=row, column=1, value="3. 'raw_product' = exatamente o que está escrito na etiqueta/prateleira.")
    row += 1
    ws.cell(row=row, column=1, value="4. 'raw_unit' = ex: '1kg', '500g', 'cx 12x395g', 'lata 1kg'.")
    row += 1
    ws.cell(row=row, column=1, value="5. 'raw_price' = número apenas (ex: 42.90).")
    row += 1
    ws.cell(row=row, column=1, value="6. 'Data' = formato AAAA-MM-DD (ex: 2026-07-15).")
    row += 1
    ws.cell(row=row, column=1, value="7. Linhas em branco são ignoradas pelo importador.")

    row += 2
    ws.cell(row=row, column=1, value="LOJAS PARA VISITAR:").font = Font(bold=True, size=12)
    row += 1
    for store in STORES_SP:
        ws.cell(row=row, column=1, value=f"- {store}")
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="INGREDIENTES MONITORADOS:").font = Font(bold=True, size=12)
    row += 1
    for ing in INGREDIENTS:
        ws.cell(row=row, column=1, value=f"- {ing}")
        row += 1

    wb.save(output_path)
    print(f"Template gerado: {output_path}")
    print(f"Colunas: {[h[1] for h in HEADERS]}")


if __name__ == "__main__":
    output = os.path.join(os.path.dirname(__file__), "..", "data", "template_visita_sp.xlsx")
    generate_template(output)
